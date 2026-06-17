"""Comandos import_projects / import_memberships contra un xlsx generado."""

from datetime import datetime

import openpyxl
import pytest
from django.contrib.auth import get_user_model
from django.core.management import call_command

from apps.catalog.models import Project, ProjectMembership

User = get_user_model()


def _build_workbook(path):
    wb = openpyxl.Workbook()
    hc = wb.active
    hc.title = "HC Total Nov 2024-2026"
    hc.append([
        "No", "NOMBRE COMPLETO", "Nombre corto", "Correo", "Área", "Puesto",
        "Cliente", "Proyecto 1", "Evaluador 1", "Evaluador 1 A", "Proyecto 2", "Evaluador 2",
    ])
    hc.append([
        1, "ANA LOPEZ PEREZ", "Ana Lopez", "ana@arena-analytics.com",
        "Analítica", "Ing Jr", "Cliente X", "Weather", "Otro Evaluador", None, None, None,
    ])

    duenos = wb.create_sheet("Proyectos Dueños")
    duenos.append([
        "ID", "Nombre", "Cliente", "Owner", "Responsable",
        "Kick-off", "Target Cierre", "Status", "Descripción",
    ])
    duenos.append([
        "P8", "Weather", "Cliente X", "Ana Lopez", "Ana Lopez",
        datetime(2026, 1, 20), datetime(2026, 6, 5), "On track", "",
    ])
    duenos.append([
        "P11", "Coppel Portal (Sistema de Gestión de Categorías)", "Coppel",
        "Carolina Palacio", "Ana Lopez",
        datetime(2025, 1, 13), datetime(2026, 5, 15), "Delayed", "",
    ])

    proj = wb.create_sheet("Proyectos")
    proj.append(["Employee", "Project", "Min of Start", "Max of End"])
    proj.append(["Ana Lopez", "Weather", datetime(2026, 1, 20), datetime(2026, 6, 5)])

    wb.save(path)


@pytest.fixture
def xlsx(tmp_path):
    path = tmp_path / "datos.xlsx"
    _build_workbook(path)
    return str(path)


@pytest.fixture
def ana(db):
    return User.objects.create_user(
        email="ana@arena-analytics.com", password="x", full_name="Ana Lopez Perez",
    )


@pytest.mark.django_db
def test_import_projects_crea_proyectos_y_usuario_faltante(xlsx, ana):
    call_command("import_projects", "--path", xlsx)

    weather = Project.objects.get(name="Weather")
    assert weather.owner == ana
    assert weather.responsable == ana
    assert weather.client == "Cliente X"
    assert weather.duration_type == Project.Duration.FINITO
    assert weather.status == Project.Status.ON_TRACK

    coppel = Project.objects.get(name__startswith="Coppel Portal")
    assert coppel.owner.email == "cpalacio@arena-analytics.com"
    assert coppel.status == Project.Status.DELAYED


@pytest.mark.django_db
def test_import_projects_es_idempotente(xlsx, ana):
    call_command("import_projects", "--path", xlsx)
    call_command("import_projects", "--path", xlsx)
    assert Project.objects.filter(name="Weather").count() == 1


@pytest.mark.django_db
def test_import_projects_dry_run_no_escribe(xlsx, ana):
    call_command("import_projects", "--path", xlsx, "--dry-run")
    assert Project.objects.count() == 0


@pytest.mark.django_db
def test_import_memberships_crea_membresia_desde_hc_total(xlsx, ana):
    call_command("import_projects", "--path", xlsx)
    call_command("import_memberships", "--path", xlsx)

    weather = Project.objects.get(name="Weather")
    assert ProjectMembership.objects.filter(project=weather, user=ana).exists()


@pytest.mark.django_db
def test_import_memberships_ignora_evaluadores(xlsx, ana):
    """Los nombres en columnas 'Evaluador N' no se crean como miembros."""
    call_command("import_projects", "--path", xlsx)
    call_command("import_memberships", "--path", xlsx)

    weather = Project.objects.get(name="Weather")
    # Solo Ana debería estar; "Otro Evaluador" del xlsx no se importa
    assert ProjectMembership.objects.filter(project=weather).count() == 1


@pytest.mark.django_db
def test_import_memberships_idempotente(xlsx, ana):
    call_command("import_projects", "--path", xlsx)
    call_command("import_memberships", "--path", xlsx)
    call_command("import_memberships", "--path", xlsx)
    weather = Project.objects.get(name="Weather")
    assert ProjectMembership.objects.filter(project=weather).count() == 1


@pytest.mark.django_db
def test_import_memberships_elimina_extras(xlsx, ana):
    """Membresías en BD que ya no están en HC Total se eliminan (sync)."""
    call_command("import_projects", "--path", xlsx)
    weather = Project.objects.get(name="Weather")
    extra = User.objects.create_user(
        email="extra@arena-analytics.com", password="x", full_name="Extra User"
    )
    ProjectMembership.objects.create(project=weather, user=extra)

    call_command("import_memberships", "--path", xlsx)

    assert ProjectMembership.objects.filter(project=weather, user=ana).exists()
    assert not ProjectMembership.objects.filter(project=weather, user=extra).exists()
