"""Importa proyectos desde la hoja 'Proyectos Dueños' del xlsx de Talento.

Idempotente por nombre de proyecto. Crea los usuarios faltantes definidos en
imports.USERS_TO_CREATE. Asigna duration_type según imports.DURATION_BY_PROJECT.

Uso (contra DATABASE_URL):
  manage.py import_projects --dry-run
  manage.py import_projects
"""

from pathlib import Path

from django.conf import settings
from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand
from django.db import transaction

from apps.catalog.models import Project
from apps.core.services import imports
from apps.core.text import normalize_name

User = get_user_model()

DEFAULT_XLSX = "Quien evalua a quien Analítica 1er S 2026.xlsx"
HC_SHEET = "HC Total Nov 2024-2026"
OWNERS_SHEET = "Proyectos Dueños"


def _header_index(row):
    return {str(v).strip().lower(): i for i, v in enumerate(row) if v is not None}


class Command(BaseCommand):
    help = "Importa proyectos desde la hoja 'Proyectos Dueños' del xlsx de Talento."

    def add_arguments(self, parser):
        parser.add_argument("--path", default=None)
        parser.add_argument("--password", default="Arena2026!")
        parser.add_argument("--dry-run", action="store_true")

    def handle(self, *args, **options):
        import openpyxl

        path = Path(options["path"]) if options["path"] else (
            Path(settings.BASE_DIR) / "docs" / "Modelos" / DEFAULT_XLSX
        )
        if not path.exists():
            self.stderr.write(self.style.ERROR(f"No se encontró {path}"))
            return

        wb = openpyxl.load_workbook(path, data_only=True)
        alias_pairs = self._alias_pairs(wb)
        dry = options["dry_run"]
        password = options["password"]

        rows = list(wb[OWNERS_SHEET].iter_rows(values_only=True))
        header = _header_index(rows[0])
        col = lambda name: header.get(name)

        created = updated = 0
        unmatched = []

        with transaction.atomic():
            index = imports.build_user_index(User.objects.all(), alias_pairs)
            for raw in rows[1:]:
                name = (raw[col("nombre")] or "").strip() if col("nombre") is not None else ""
                if not name:
                    continue
                owner_name = raw[col("owner")] if col("owner") is not None else None
                resp_name = raw[col("responsable")] if col("responsable") is not None else None

                lead, lead_action = imports.resolve_or_create_user(
                    owner_name, index, password=password, dry=dry
                )
                if lead_action in ("would_create", "created"):
                    self.stdout.write(f"[{'dry' if dry else 'ok'}] usuario owner: {owner_name}")
                if lead is None:
                    unmatched.append(f"Owner no resuelto: {owner_name!r} ({name})")
                    continue
                responsable, _ = imports.resolve_or_create_user(
                    resp_name, index, password=password, dry=dry
                )

                client = (raw[col("cliente")] or "").strip() if col("cliente") is not None else ""
                status_raw = (raw[col("status")] or "").strip().lower() if col("status") is not None else ""
                status = Project.Status.DELAYED if "delay" in status_raw else Project.Status.ON_TRACK
                kickoff = imports.to_date(raw[col("kick-off")]) if col("kick-off") is not None else None
                target = imports.to_date(raw[col("target cierre")]) if col("target cierre") is not None else None
                duration = imports.DURATION_BY_PROJECT.get(normalize_name(name), Project.Duration.FINITO)

                if dry:
                    self.stdout.write(
                        f"[dry] {name} | lead={lead} | resp={responsable} | "
                        f"{duration} | {status} | {kickoff}–{target}"
                    )
                    continue

                project, is_new = Project.objects.get_or_create(
                    name=name, defaults={"lead": lead},
                )
                project.client = client
                project.lead = lead
                project.responsable = responsable
                project.kickoff = kickoff
                project.target_close = target
                project.status = status
                project.duration_type = duration
                project.is_active = True
                project.save()
                created += int(is_new)
                updated += int(not is_new)

            if dry:
                transaction.set_rollback(True)

        for msg in unmatched:
            self.stdout.write(self.style.WARNING(msg))
        self.stdout.write(self.style.SUCCESS(
            f"Proyectos — nuevos: {created} · actualizados: {updated} · sin resolver: {len(unmatched)}"
        ))

    def _alias_pairs(self, wb):
        """(nombre corto, correo) desde HC Total para resolver por correo."""
        if HC_SHEET not in wb.sheetnames:
            return []
        rows = list(wb[HC_SHEET].iter_rows(values_only=True))
        if not rows:
            return []
        header = _header_index(rows[0])
        i_short = header.get("nombre corto")
        i_email = header.get("correo")
        if i_short is None or i_email is None:
            return []
        pairs = []
        for raw in rows[1:]:
            short = raw[i_short] if i_short < len(raw) else None
            email = raw[i_email] if i_email < len(raw) else None
            if short and email:
                pairs.append((str(short).strip(), str(email).strip()))
        return pairs
