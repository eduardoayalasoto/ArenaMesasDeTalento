"""Importa y sincroniza membresías desde la hoja 'HC Total' del xlsx.

Cada fila es una persona. Las columnas 'Proyecto N' enumeran los proyectos a
los que pertenece. Las columnas 'Evaluador N' se ignoran. Para cada proyecto
del catálogo, crea membresías faltantes y elimina las que ya no están en el
xlsx. Idempotente. Soporta --dry-run.

Uso:
  manage.py import_memberships --dry-run
  manage.py import_memberships
"""

from pathlib import Path

from django.conf import settings
from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand
from django.db import transaction

from apps.catalog.models import Project, ProjectMembership

User = get_user_model()

DEFAULT_XLSX = "Quien evalua a quien Analítica 1er S 2026.xlsx"
HC_SHEET = "HC Total Nov 2024-2026"


class Command(BaseCommand):
    help = "Importa/sincroniza membresías desde la hoja 'HC Total' del xlsx de Talento."

    def add_arguments(self, parser):
        parser.add_argument("--path", default=None)
        parser.add_argument("--dry-run", action="store_true")

    def handle(self, *args, **options):
        import openpyxl

        path = Path(options["path"]) if options["path"] else (
            Path(settings.BASE_DIR) / "docs" / "Modelos" / DEFAULT_XLSX
        )
        if not path.exists():
            self.stderr.write(self.style.ERROR(f"No se encontró {path}"))
            return

        dry = options["dry_run"]
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[HC_SHEET]

        headers = [c.value for c in ws[1]]
        proj_col_indices = [
            i for i, h in enumerate(headers)
            if h and str(h).startswith("Proyecto")
        ]
        email_col = next(
            (i for i, h in enumerate(headers) if h and "correo" in str(h).lower()),
            None,
        )
        if email_col is None:
            self.stderr.write(self.style.ERROR("No se encontró columna 'Correo' en HC Total"))
            return

        db_projects = {p.name: p for p in Project.objects.all()}

        # desired: project_name -> {user_pk, ...}
        desired: dict[str, set] = {}
        unresolved_users: list[str] = []
        unresolved_projects: list[str] = []

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row[1]:
                continue
            raw_email = row[email_col] if email_col < len(row) else None
            if not raw_email:
                continue
            email = str(raw_email).strip()

            try:
                user = User.objects.get(email=email)
            except User.DoesNotExist:
                unresolved_users.append(f"r{i}: {row[1]} ({email!r})")
                continue

            for ci in proj_col_indices:
                if ci >= len(row) or not row[ci]:
                    continue
                proj_name = str(row[ci]).strip()
                if proj_name not in db_projects:
                    if proj_name not in unresolved_projects:
                        unresolved_projects.append(proj_name)
                    continue
                desired.setdefault(proj_name, set()).add(user.pk)

        created = deleted = kept = 0

        with transaction.atomic():
            for proj_name, user_pks in desired.items():
                project = db_projects[proj_name]

                for pk in user_pks:
                    user = User.objects.get(pk=pk)
                    _, is_new = ProjectMembership.objects.get_or_create(
                        project=project, user=user,
                    )
                    if is_new:
                        created += 1
                        self.stdout.write(f"  + {proj_name}: {user.full_name}")
                    else:
                        kept += 1

                extra = project.memberships.select_related("user").exclude(
                    user__pk__in=user_pks
                )
                for m in extra:
                    self.stdout.write(f"  - {proj_name}: {m.user.full_name}")
                    deleted += 1
                    if not dry:
                        m.delete()

            if dry:
                transaction.set_rollback(True)

        for msg in unresolved_users:
            self.stdout.write(self.style.WARNING(f"Usuario sin match: {msg}"))
        for proj in unresolved_projects:
            self.stdout.write(self.style.WARNING(f"Proyecto sin match en BD: {proj!r}"))

        verb = "[dry] " if dry else ""
        self.stdout.write(self.style.SUCCESS(
            f"{verb}Membresías — creadas: {created} · eliminadas: {deleted}"
            f" · sin cambio: {kept} · sin resolver: {len(unresolved_users)}"
        ))
